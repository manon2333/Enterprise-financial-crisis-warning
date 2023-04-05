"""
功能：
依据191家ST企业的第一次ST年份时的资产规模和所在行业，进行1：5配对
注意：
偶尔会存在ST企业，与它相匹配的企业少于5家。这种情况下，需要手动寻找并填写
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill


def custom_sort(filtered_result, st_total_assets):
    # 计算 TotalAssets 列中每个元素与 st_total_assets 的差的绝对值
    abs_diff = abs(filtered_result['TotalAssets'] - st_total_assets)
    filtered_result['AbsDiff'] = abs_diff
    sorted_filtered_result = filtered_result.sort_values(by='AbsDiff')  # 按照 AbsDiff 列升序排列
    sorted_filtered_result = sorted_filtered_result.drop(columns='AbsDiff')  # 删除添加的 AbsDiff 列
    return sorted_filtered_result


def write_selected_elements_to_excel(book, selected_elements, start_row, st_stock_code, first_st_year, st_total_assets):
    """
    将DataFrame类型的数据写入Excel文件中的paired_results表中，从指定行开始写入
    :param book: Workbook, 要写入的Excel文件对象
    :param selected_elements: 要写入Excel文件中的数据
    :param start_row: 写入数据的起始行
    """
    # for index, row in selected_elements.iterrows():
    #     writer = pd.ExcelWriter('D:\\Enterprise financial crisis warning\\test.xlsx', engine='xlsxwriter')
    #     row.to_excel(writer, sheet_name='paired_results', index=False)
    #     writer.save()
    ws = book['paired_results']
    # 写入st股票代码信息
    st_information = pd.DataFrame({'Symbol': [st_stock_code],
                                   'FirstSTDate': [first_st_year],
                                   'ST_TotalAssets': [st_total_assets]})
    for _, row in st_information.iterrows():
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=start_row - 1, column=col_num, value=value)
    # 将st股票代码设置为黄色填充样式，方便区分
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws[start_row - 1]:
        cell.fill = fill
    # 写入与st股票代码相匹配的股票代码信息（可能少于5个）
    for _, row in selected_elements.iterrows():
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=start_row, column=col_num, value=value)
        start_row += 1
    book.save('D:\\Enterprise financial crisis warning\\paired_results.xlsx')


if __name__ == '__main__':
    # 初始化
    filtered_stock_codes = []  # 初筛后，满足条件的股票代码
    filtered_information = []
    data_file_path = 'D:\\Enterprise financial crisis warning\\data.xlsx'
    paired_result_file_path = 'D:\\Enterprise financial crisis warning\\paired_results.xlsx'
    start_row = 3  # 从paired_results.xlsx中，名为paired_results的sheet的第3行开始写入数据

    # 新建保存最终结果的Excel文件
    book = openpyxl.Workbook()
    book.create_sheet('paired_results')
    book.save(paired_result_file_path)
    book = openpyxl.load_workbook(paired_result_file_path)
    final_paired_result = book['paired_results']
    header = ['Symbol', 'FirstSTDate', 'ST_TotalAssets']  # 列名
    final_paired_result.append(header)
    book.save(paired_result_file_path)

    # 读取Excel文件
    sheet1 = pd.read_excel(data_file_path, sheet_name='Non_ST_stock_codes', dtype={'Symbol': str})
    sheet2 = pd.read_excel(data_file_path, sheet_name='ST_stock_codes', dtype={'Symbol': str})
    sheet3 = pd.read_excel(data_file_path, sheet_name='Non_ST_asset_size', dtype={'Symbol': str})
    sheet3['EndDate'] = sheet3['EndDate'].str.slice(stop=4)
    sheet3['EndDate'] = sheet3['EndDate'].astype(str)

    for index2, row2 in sheet2.iterrows():
        st_stock_code = row2['Symbol']
        st_industry_code = row2['IndustryCode']
        first_st_year = row2['FirstSTDate'][:4]
        print("---------------------------------以下是与股票代码为{0}相匹配的股票代码信息---------------------------------------".format(st_stock_code))
        print("该股票代码的first_st_year是{0}".format(first_st_year))
        st_total_assets = row2['ST_TotalAssets']
        # 在sheet1中寻找属于同一个行业的股票代码
        filtered_data = sheet1[sheet1['IndustryCode'] == st_industry_code]
        filtered_stock_codes = filtered_data['Symbol'].tolist()
        # 筛去filtered_stock_codes中，数值等于当前st_stock_code的字符串
        filtered_stock_codes = np.array(filtered_stock_codes)
        mask = filtered_stock_codes != st_stock_code  # 匹配条件
        filtered_stock_codes = filtered_stock_codes[mask].tolist()
        # 在sheet3中寻找对应股票代码，在first_ST_year的TotalAssets
        for code in filtered_stock_codes:
            mask = (sheet3['Symbol'] == code) & (sheet3['EndDate'] == first_st_year) & (sheet3['StateType'] == 'A')   # 匹配条件
            temp = sheet3.loc[mask]
            # print(temp)
            filtered_information.append(temp)
        filtered_result = pd.concat(filtered_information, axis=0, ignore_index=True)  # 将所有数据拼接在一起
        # print(filtered_result)
        if filtered_result.empty:
            print("与股票代码为{0}相匹配的股票代码不存在".format(st_stock_code))
        row = filtered_result.shape[0]
        if row <= 5:
            selected_elements = filtered_result[['Symbol', 'EndDate', 'TotalAssets']]
        else:
            sorted_filtered_result = custom_sort(filtered_result, st_total_assets)
            selected_elements = sorted_filtered_result.filter(['Symbol', 'EndDate', 'TotalAssets']).head(5)
        print("以下是与该股票代码相匹配（同行业，st时期资产规模最接近）的股票代码：")
        print(selected_elements)
        # 将筛选出的数据写入 Excel 表格中
        book = openpyxl.load_workbook(paired_result_file_path)
        write_selected_elements_to_excel(book, selected_elements, start_row, st_stock_code, first_st_year, st_total_assets)
        print("与股票代码为{0}相匹配的数据已成功写入paired_results中的第{1}-第{2}行".format(st_stock_code, start_row, start_row+4))
        start_row = start_row + 6
        filtered_information[:] = []
        filtered_result = pd.DataFrame()
    book.close()










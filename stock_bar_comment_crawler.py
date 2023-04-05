import requests
from lxml import etree
import xlwt as wt
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime


# 初始化
save_path = 'D:\\Enterprise financial crisis warning\\stock_bar_comment_information.xlsx'
header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36 Edg/98.0.1108.62"
}
title = []
author = []
time = []
comments_num = []
views = []
comments_address = []
span = 8  # 时间跨度


def locate_startpage_and_endpage(code, start_date, end_date):
    is_first_iteration = True
    is_last_iteration = True
    start_page = 0
    end_page = 0

    # 获取对应股票代码的股吧总页数
    # 因总页数采用JS动态渲染，故只能用 Selenium 来获取
    options = webdriver.ChromeOptions().add_argument('headless')
    browser = webdriver.Chrome(options=options)
    browser.get('http://guba.eastmoney.com/list,{0},f.html'.format(code))  # 以发帖时间排序
    browser.implicitly_wait(2)
    total_page = browser.find_element(By.XPATH, '//*[@id="articlelistnew"]/div[82]/span/span/span[1]/span')
    total_page = int(total_page.text)

    for page in range(1, total_page + 1):
        # 获得当前页下，第1条评论的详情页链接
        url = "http://guba.eastmoney.com/list,{0},f_{1}.html".format(code, page)
        response = requests.get(url=url, headers=header)
        response.content.decode("utf-8")
        a = etree.HTML(response.text)
        first_comment_url_div = a.xpath('//*[@id="articlelistnew"]/div[2]/span[3]/a')
        # print(first_comment_url_div[0].get('href'))
        first_comment_href = 'https://guba.eastmoney.com' + first_comment_url_div[0].get('href')

        # 获得第1条评论的发布年份
        sub_page_response = requests.get(url=first_comment_href, headers=header)
        a = etree.HTML(sub_page_response.text)

        # 确定 start_page 和 end_page
        try:
            post_year_div = a.xpath('//*[@id="zwconttb"]/div[2]')
            temp = post_year_div[0].text
            post_year = temp.strip().split()[1]
            post_year_dt = datetime.strptime(post_year, '%Y-%m-%d')
            start_date_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d')

            # 条件判断
            if (page == total_page) & (post_year_dt > end_date_dt):
                print("股票代码为{0}在对应时间段内，无法找到任何股吧评论".format(code))
            elif (page == 1) & (post_year_dt < start_date_dt):
                print("股票代码为{0}在对应时间段内，无法找到任何股吧评论".format(code))
                break
            else:
                if (post_year_dt < end_date_dt) & (is_first_iteration == True):
                    start_page = page
                    is_first_iteration = False
                elif (post_year_dt < start_date_dt) & (is_last_iteration == True):
                    end_page = page - 1
                    is_last_iteration = False
        except:
            print("#######出现错误：该帖子的类型是“专栏正文”，而不是“正文”")
            pass

    is_first_iteration = True
    is_last_iteration = True
    return start_page, end_page


def obtain_stock_comments(code, start_page, end_page, work_book):
    url = "http://guba.eastmoney.com/list,{0},f_{1}.html".format(code, start_page)
    # http://guba.eastmoney.com/list,000004,f_2.html
    start_page = start_page + 1
    response = requests.get(url, headers=header)
    response.content.decode("utf-8")
    a = etree.HTML(response.text)
    div_list = a.xpath('//*[@id="articlelistnew"]/div')
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[3]/a/@title'))
        if len(i) != 0:
            title.append(i)
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[3]/a/@href'))
        if len(i) != 0:
            comments_address.append(i)
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[4]/a//text()'))
        if len(i) != 0:
            author.append(i)
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[5]/text()'))
        if len(i) != 0:
            time.append(i)
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[1]/text()'))
        if len(i) != 0:
            views.append(i)
    for i in div_list[1:]:
        i = ''.join(i.xpath('./span[2]/text()'))
        if len(i) != 0:
            comments_num.append(i)
            
    # 结果存入excel表中
    work_book.create_sheet('{0}'.format(code))
    work_book = openpyxl.load_workbook(save_path)
    work_sheet = work_book['{0}'.format(code)]
    headers = ['标题', '作者', '时间', '阅读量', '评论数']
    work_sheet.append(headers)
    for c in range(80 * (end_page - start_page + 1)):
        work_sheet.write(c + 1, 0, title[c])
        work_sheet.write(c + 1, 2, time[c])
        work_sheet.write(c + 1, 1, author[c])
        work_sheet.write(c + 1, 3, views[c])
        work_sheet.write(c + 1, 4, comments_num[c])


if __name__ == '__main__':
    # 新建保存最终结果的Excel文件
    work_book = openpyxl.Workbook()
    work_book.save(save_path)

    # 获取股票代码
    df = pd.read_excel('D:\\Enterprise financial crisis warning\\paired_result.xlsx',
                       sheet_name='paired_results', dtype={'Symbol': str, 'FirstSTDate': int})
    deduplicated_data = df.drop_duplicates(subset=['Symbol'], keep='first').reset_index(drop=True)
    code_list = deduplicated_data['Symbol'].tolist()

    # 爬取对应股票代码的股吧评论
    for code in code_list:
        print('----------------------------------------------------------------------------')
        print('即将爬取的是股票代码为{0}的股吧评论'.format(code))
        start_date = str(deduplicated_data[deduplicated_data['Symbol'] == code]['FirstSTDate'].iloc[0] - span) + '-01-01'
        end_date = str(deduplicated_data[deduplicated_data['Symbol'] == code]['FirstSTDate'].iloc[0] + 1) + '-01-01'
        print('时间范围是：{0} —— {1}'.format(start_date, end_date))
        start_page, end_page = locate_startpage_and_endpage(code, start_date, end_date)
        print('起始页：{0}；终止页：{1}'.format(start_page, end_page))
        obtain_stock_comments(code, start_page, end_page, work_book)



